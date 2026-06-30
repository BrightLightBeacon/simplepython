import os
import sys
import subprocess
import importlib

def install_and_import(module_name, package_name=None):
    if package_name is None:
        package_name = module_name
    try:
        importlib.import_module(module_name)
    except ImportError:
        print(f"Missing required dependency '{package_name}'. Attempting to install automatically...", flush=True)
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
            importlib.import_module(module_name)
            print(f"Successfully installed {package_name}!", flush=True)
        except Exception as e:
            print(f"Error installing {package_name} automatically: {e}", file=sys.stderr, flush=True)
            print(f"Please install it manually using: pip install {package_name}", file=sys.stderr, flush=True)
            sys.exit(1)

install_and_import("pandas")
install_and_import("openpyxl")

import argparse
import datetime
import calendar
import re
import json
import glob
from pathlib import Path
import pandas as pd
import openpyxl
from copy import copy
from openpyxl.worksheet.formula import ArrayFormula

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)

def find_template_file(filename, fallback_dir="шаблони"):
    for path in Path(ROOT_DIR).rglob(filename):
        parts = [p.lower() for p in path.parts]
        if path.is_file() and not any(p.startswith('.') or p in ('venv', 'env', 'build', 'dist', 'node_modules', '__pycache__') for p in parts):
            return str(path)
    return os.path.join(ROOT_DIR, fallback_dir, filename)

def get_month_name_ukr(month_idx):
    names = {
        1: "січень", 2: "лютий", 3: "березень", 4: "квітень",
        5: "травень", 6: "червень", 7: "липень", 8: "серпень",
        9: "вересень", 10: "жовтень", 11: "листопад", 12: "грудень"
    }
    return names.get(month_idx, "")

def update_formula_rows(formula_text, src_start_row, tgt_start_row):
    if not isinstance(formula_text, str):
        return formula_text
    
    if "ROW()-13" in formula_text:
        return formula_text
        
    new_text = formula_text
    for offset in [3, 2, 1, 0]:
        src_row = src_start_row + offset
        tgt_row = tgt_start_row + offset
        pattern = rf"(?<!\d){src_row}(?!\d)"
        new_text = re.sub(pattern, str(tgt_row), new_text)
        
    return new_text

def copy_row_style(ws, src_row, tgt_row, src_start_row, tgt_start_row):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        tgt_cell = ws.cell(row=tgt_row, column=col)
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)
        
        if isinstance(src_cell.value, ArrayFormula):
            old_ref = src_cell.value.ref
            old_text = src_cell.value.text
            new_ref = update_formula_rows(old_ref, src_start_row, tgt_start_row)
            new_text = update_formula_rows(old_text, src_start_row, tgt_start_row)
            tgt_cell.value = ArrayFormula(new_ref, new_text)
        elif isinstance(src_cell.value, str) and src_cell.value.startswith('='):
            tgt_cell.value = update_formula_rows(src_cell.value, src_start_row, tgt_start_row)
        else:
            tgt_cell.value = src_cell.value

def is_service_agreement(param_text):
    """Check if the parameter contains service agreement pattern ${sa;company;money;timestamp}"""
    if not param_text or pd.isna(param_text):
        return False
    
    param_str = str(param_text).strip()
    # Pattern to match ${sa;company;money;timestamp}
    pattern = r'\$\{sa;[^;]+;[^;]+;[^}]+\}'
    return bool(re.search(pattern, param_str))

def is_exclude_param(param_text):
    """Check whether the parameter cell contains ${exclude}."""
    if not param_text or pd.isna(param_text):
        return False
    s = str(param_text)
    return bool(re.search(r'\$\{exclude\}', s, re.IGNORECASE))

def extract_driver_from_parameter(param_text):
    """
    Extract driver name from parameter if it contains водій {Name} pattern (or legacy ${driver;Name}).
    Example: водій {Бондаренко Сергій Валерійович} -> returns 'Бондаренко Сергій Валерійович'
    Returns None if pattern not found.
    """
    if not param_text or pd.isna(param_text):
        return None
    
    param_str = str(param_text).strip()
    
    # Pattern to match водій {Name}
    pattern_new = r'водій\s*\{([^}]+)\}'
    match_new = re.search(pattern_new, param_str, re.IGNORECASE)
    if match_new:
        driver_name = match_new.group(1).strip()
        return driver_name if driver_name else None
        
    # Legacy pattern to match ${driver;Name}
    pattern_old = r'\$\{driver;([^}]+)\}'
    match_old = re.search(pattern_old, param_str)
    if match_old:
        driver_name = match_old.group(1).strip()
        return driver_name if driver_name else None
    
    return None

def format_driver_initials(full_name):
    """Formats driver full name to 'Surname\nF. P.' or 'Surname\nF.'."""
    if not isinstance(full_name, str) or not full_name.strip():
        return "NoName"
        
    parts = [p.strip() for p in full_name.split() if p.strip()]
    if len(parts) == 0:
        return "NoName"
    if len(parts) == 1:
        return parts[0]
        
    first_names = {
        "олександр", "сергій", "дмитро", "андрій", "володимир", "віталій", "василь", 
        "юрій", "ігор", "олег", "олексій", "роман", "михайло", "микола", "ярослав", 
        "артем", "тарас", "іван", "богдан", "павло", "євген", "анатолій", "віктор", 
        "валерій", "руслан", "владислав", "денис", "максим", "вадим", "петро", 
        "костянтин", "антон", "олександр", "в'ячеслав", "вячеслав", "геннадій", 
        "леонід", "борис", "григорій", "микита", "назар", "захар", "кирило",
        "анна", "ольга", "тетяна", "олена", "ірина", "наталія", "натрія", "марія", 
        "світлана", "катерина", "людмила", "оксана", "галина", "надія", "любов", 
        "валентина", "юлія", "лариса", "олександра", "євгенія", "вікторія", "марина", 
        "дарія", "дар'я", "дар’я", "аліна", "анастасія", "надія", "людмила", "ніна"
    }
    
    surname_suffixes = (
        "енко", "ук", "юк", "ич", "ов", "ев", "єв", "ський", 
        "цький", "ська", "цька", "шин", "ин", "ий", "их"
    )
    
    def is_patronymic(word):
        w = word.lower()
        return (w.endswith("ович") or w.endswith("евич") or w.endswith("євич") or 
                w.endswith("івна") or w.endswith("ївна") or w.endswith("евна"))

    if len(parts) >= 3:
        pat_index = -1
        for idx, part in enumerate(parts):
            if is_patronymic(part):
                pat_index = idx
                break
                
        if pat_index != -1:
            patronymic = parts[pat_index]
            if pat_index == 1:
                firstname = parts[0]
                surname = parts[2]
            elif pat_index == 2:
                surname = parts[0]
                firstname = parts[1]
            else:
                surname = parts[0]
                firstname = parts[1]
        else:
            surname = parts[0]
            firstname = parts[1]
            patronymic = parts[2]
            
        f_init = f"{firstname[0]}." if firstname else ""
        p_init = f"{patronymic[0]}." if patronymic else ""
        initials_str = f"{f_init} {p_init}".strip()
        return f"{surname}\n{initials_str}"
        
    elif len(parts) == 2:
        w1, w2 = parts[0], parts[1]
        w1_lower, w2_lower = w1.lower(), w2.lower()
        
        w1_is_first = w1_lower in first_names
        w2_is_first = w2_lower in first_names
        
        if w1_is_first and not w2_is_first:
            firstname, surname = w1, w2
        elif w2_is_first and not w1_is_first:
            surname, firstname = w1, w2
        else:
            w1_is_surname = w1_lower.endswith(surname_suffixes)
            w2_is_surname = w2_lower.endswith(surname_suffixes)
            
            if w1_is_surname and not w2_is_surname:
                surname, firstname = w1, w2
            elif w2_is_surname and not w1_is_surname:
                firstname, surname = w1, w2
            else:
                surname, firstname = w1, w2
                
        f_init = f"{firstname[0]}." if firstname else ""
        return f"{surname}\n{f_init}"

def normalize_name(name):
    """Normalize names for matching by converting to lowercase, stripping whitespace, and standardizing apostrophes."""
    if not isinstance(name, str):
        return ""
    s = name.strip().lower()
    s = re.sub(r"[’'‘`]", "'", s)
    s = re.sub(r"\s+", " ", s)
    return s

def parse_date_range(range_str):
    """Parse date range string in 'DD.MM.YYYY - DD.MM.YYYY' format to start and end dates."""
    try:
        parts = range_str.split("-")
        if len(parts) == 2:
            start_str = parts[0].strip()
            end_str = parts[1].strip()
            start_date = datetime.datetime.strptime(start_str, "%d.%m.%Y").date()
            end_date = datetime.datetime.strptime(end_str, "%d.%m.%Y").date()
            return start_date, end_date
    except Exception as e:
        pass
    return None

def matches_company(vac_company, target_company):
    if not vac_company or vac_company.strip() in ("", "*"):
        return True
    vc = vac_company.strip().upper()
    tc = target_company.strip().upper()
    return vc in tc or tc in vc

def main():
    parser = argparse.ArgumentParser(description="Generate timesheets based on Excel data.")
    parser.add_argument("-m", "--month", type=str, help="Target month in mmyy format (e.g. 0526). Defaults to the previous month if omitted.")
    parser.add_argument("-c", "--company", type=str, help="Specific company to process ('zia' for ЗІАВТОТРАНС, 'zet' for ЗЕТТРА). Defaults to both if omitted.")
    parser.add_argument("-f", "--file", type=str, help="Specific source Excel file path.")
    
    args = parser.parse_args()
    
    # Load and parse vacations from відпустки.txt inside ROOT_DIR
    vacations_path = os.path.join(ROOT_DIR, "відпустки.txt")
    if not os.path.exists(vacations_path):
        fallback_vac = os.path.join(ROOT_DIR, "відпустки")
        if os.path.exists(fallback_vac):
            vacations_path = fallback_vac

    parsed_vacations = []
    if os.path.exists(vacations_path):
        try:
            with open(vacations_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = [p.strip() for p in line.split(",")]
                    if len(parts) >= 4:
                        comp_str = parts[0]
                        v_name = parts[1]
                        v_range = parts[2]
                        v_type = ",".join(parts[3:]).strip()
                    elif len(parts) == 3:
                        comp_str = ""
                        v_name = parts[0]
                        v_range = parts[1]
                        v_type = parts[2]
                    else:
                        continue
                        
                    if v_name and v_range and v_type:
                        norm_name = normalize_name(v_name)
                        date_bounds = parse_date_range(v_range)
                        if date_bounds:
                            start_date, end_date = date_bounds
                            parsed_vacations.append({
                                "company": comp_str,
                                "name": norm_name,
                                "start": start_date,
                                "end": end_date,
                                "type": v_type
                            })
            print(f"Loaded {len(parsed_vacations)} valid vacation records from {vacations_path}")
        except Exception as e:
            print(f"Error loading or parsing vacations file: {e}")
    else:
        print(f"Warning: Vacations file not found at {vacations_path}")

    # Dynamic template resolution
    template_path = find_template_file("template_timesheet.xlsx")
    if not os.path.exists(template_path):
        for candidate_name in ["шаблон-табелю.xlsx", "шаблон_табелю.xlsx", "шаблон-табель.xlsx"]:
            candidate = find_template_file(candidate_name)
            if os.path.exists(candidate):
                template_path = candidate
                break

    output_dir = os.path.join(ROOT_DIR, "табелі")
    if os.path.exists(os.path.join(ROOT_DIR, "табель")):
        output_dir = os.path.join(ROOT_DIR, "табель")
    os.makedirs(output_dir, exist_ok=True)
    
    mmyy = args.month
    excel_path = None
    if args.file:
        if os.path.exists(args.file):
            excel_path = args.file
        elif os.path.exists(os.path.join(ROOT_DIR, args.file)):
            excel_path = os.path.join(ROOT_DIR, args.file)

    if not excel_path and mmyy:
        candidates = glob.glob(os.path.join(ROOT_DIR, f"*{mmyy}*.xls*"))
        candidates = [f for f in candidates if not os.path.basename(f).startswith('~$')]
        if candidates:
            excel_path = candidates[0]

    if not excel_path:
        for name_pattern in ["Дебетовий список*.xls*", "дебеторка*.xls*"]:
            candidates = glob.glob(os.path.join(ROOT_DIR, name_pattern))
            candidates = [f for f in candidates if not os.path.basename(f).startswith('~$')]
            if candidates:
                excel_path = candidates[0]
                break

    if not excel_path:
        source_files = glob.glob(os.path.join(ROOT_DIR, "*.xls*"))
        source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
        if len(source_files) == 1:
            excel_path = source_files[0]
        elif len(source_files) > 1:
            excel_path = source_files[0]
            print(f"Multiple Excel files found in {ROOT_DIR}. Using {excel_path}")

    if not excel_path or not os.path.exists(excel_path):
        print(f"Error: Target Excel workbook not found in {ROOT_DIR}.")
        return

    print(f"Reading data from {excel_path} ...")
    
    try:
        # Load the first sheet available in the workbook
        df = pd.read_excel(excel_path)
    except Exception as e:
        print(f"Error reading {excel_path}: {e}")
        return
    
    # Find column names dynamically to handle variations safely
    def find_column(options, exact=False):
        # 1. Try exact case-insensitive match first to avoid substring collision (e.g. matching 'водій' instead of a longer description)
        for opt in options:
            for col in df.columns:
                if str(col).strip().lower() == opt.lower():
                    return col
        # 2. Try substring match if exact not requested
        if not exact:
            for opt in options:
                for col in df.columns:
                    if opt.lower() in str(col).lower():
                        return col
        return None

    col_company = find_column(["товариство"])
    col_driver = find_column(["водій"])
    col_start = find_column(["початок рейсу", "навантаження дата"])
    col_end = find_column(["завершення рейсу", "розвантаження дата"])
    col_method = find_column(["спосіб"], exact=True)
    col_param = find_column(["параметр", "накази та цпх", "нотатка"])
    col_status = find_column(["статус водія"])
    
    if not all([col_company, col_driver, col_start, col_end]):
        print("Missing required columns in the Excel file.")
        print("Found columns:", df.columns.tolist())
        return

    # Normalize company names to group variations (e.g., 'ЗІА' -> 'ЗІАВТОТРАНС')
    def normalize_company(val):
        if pd.isna(val):
            return val
        val_str = str(val).strip().upper()
        if val_str in ('ЗІА', 'ЗІАВТОТРАНС'):
            return 'ЗІАВТОТРАНС'
        if val_str in ('ЗЕТ', 'ЗЕТТРА'):
            return 'ЗЕТТРА'
        return str(val).strip()
        
    df[col_company] = df[col_company].apply(normalize_company)

    # --- Filter: explicit exclude via parameter and service agreements ---
    if col_param:
        before_explicit = len(df)
        df = df[~df[col_param].apply(is_exclude_param)]
        explicit_excluded = before_explicit - len(df)
        if explicit_excluded:
            print(f"Excluded rows via ${{exclude}} marker: {explicit_excluded}")

        before_sa = len(df)
        df = df[~df[col_param].apply(is_service_agreement)]
        sa_excluded = before_sa - len(df)
        if sa_excluded:
            print(f"Excluded rows via ${{sa;...}} marker (Service Agreements): {sa_excluded}")

    # --- Apply driver name overrides (${driver;Name}) ---
    if col_param:
        overridden_drivers = df[col_param].apply(extract_driver_from_parameter)
        has_override = overridden_drivers.notna()
        if has_override.any():
            print(f"Applying driver name overrides for {has_override.sum()} rows...")
            df.loc[has_override, col_driver] = overridden_drivers[has_override]

    # Filter invalid rows (ensure company and driver exist)
    df = df.dropna(subset=[col_company, col_driver])
    
    # Filter trips to only include "безготівк. з ПДВ" in column "спосіб"
    if col_method:
        df = df[df[col_method].apply(lambda x: "безготівк. з пдв" in str(x).lower())]
    else:
        print("Warning: Column 'спосіб' not found. Cannot filter by 'безготівк. з ПДВ'.")

    # --- Filter out drivers with status "найм" ---
    # Exceptions: rows with ${assignment} in the parameter column are kept regardless
    if col_status:
        before_status_filter = len(df)
        is_najm = df[col_status].astype(str).str.upper().str.contains('НАЙМ', na=False, regex=True)
        if col_param:
            has_assignment_param = df[col_param].astype(str).str.contains(r'\$\{assignment\}', na=False, regex=True)
            najm_exception = has_assignment_param
        else:
            najm_exception = pd.Series(False, index=df.index)
        
        status_filter = ~is_najm | najm_exception
        df = df[status_filter]
        print(f"Rows after excluding drivers with status 'найм': {len(df)} (excluded {before_status_filter - len(df)})")
    
    # Determine target month
    if mmyy:
        target_month = int(mmyy[:2])
        target_year = 2000 + int(mmyy[2:])
    else:
        # Default to previous month
        today = datetime.date.today()
        first_day_of_this_month = today.replace(day=1)
        prev_month_date = first_day_of_this_month - datetime.timedelta(days=1)
        target_month = prev_month_date.month
        target_year = prev_month_date.year

    mmyy_str = mmyy if mmyy else f"{target_month:02d}{target_year % 100:02d}"
    
    # Identify companies to process and filter out garbage rows (like totals row "24")
    valid_companies = [str(c).strip() for c in df[col_company].unique() if pd.notna(c) and str(c).strip() != "" and ("ЗІА" in str(c).upper() or "ЗЕТ" in str(c).upper())]
    
    if args.company:
        comp_arg = args.company.lower()
        if comp_arg == "zia":
            target_companies = [c for c in valid_companies if "ЗІА" in c.upper()]
        elif comp_arg == "zet":
            target_companies = [c for c in valid_companies if "ЗЕТ" in c.upper()]
        else:
            print(f"Unknown company '{args.company}'. Please use 'zia' or 'zet'.")
            return
            
        if not target_companies:
            print(f"Data for company '{args.company}' not found in the sheet.")
            return
    else:
        target_companies = valid_companies

    if not target_companies:
        print("No valid companies (zia or zet) found in the data.")
        return

    # --- ANALYSIS PHASE ---
    print("\n" + "="*50)
    print("ANALYSIS SUMMARY:")
    print("="*50)
    print(f"Source Excel:       {excel_path}")
    print(f"Timesheet Template: {template_path}")
    print(f"Target Month:       {get_month_name_ukr(target_month).capitalize()} {target_year}")
    print(f"Output Directory:   {output_dir}")
    print("\nData found for the following companies:")
    
    companies_info = {}
    for company_str in target_companies:
        company_df = df[df[col_company].apply(lambda x: str(x).strip() == company_str)]
        drivers = sorted(company_df[col_driver].unique())
        trips_count = len(company_df)
        companies_info[company_str] = {
            "drivers": list(drivers),
            "trips": trips_count
        }
        print(f"  - {company_str}: {len(drivers)} unique drivers, {trips_count} total trips")
    
    print("\nFiles to be generated:")
    for company_str in target_companies:
        out_name = f"Табель обліку робочого часу {company_str} {mmyy_str}.xlsx"
        print(f"  -> {out_name}")
    print("="*50 + "\n")
    
    # Prompt user
    try:
        input("Press Enter to proceed with generation, or press Ctrl+C to abort...")
    except KeyboardInterrupt:
        print("\nOperation aborted by user.")
        return

    # --- GENERATION PHASE ---
    print("\nStarting timesheet generation...")
    for company_str in target_companies:
        print(f"Processing company: {company_str} ...")
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
        except Exception as e:
            print(f"Error loading template {template_path}: {e}")
            return
            
        director = ""
        full_company_name = company_str
        if "ЗІА" in company_str.upper():
            director = "Урда Анна Олександрівна"
            full_company_name = 'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ "ЗІАВТОТРАНС", код ЄДРПОУ 42690553'
        elif "ЗЕТ" in company_str.upper():
            director = "Рускевич Дар’я Іванівна"
            full_company_name = 'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ "ЗЕТТРА", код ЄДРПОУ 44976850'
            
        last_day = calendar.monthrange(target_year, target_month)[1]
        first_date_str = f"01.{target_month:02d}.{target_year}"
        last_date_str = f"{last_day:02d}.{target_month:02d}.{target_year}"
            
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value
                    if "${company}" in v:
                        v = v.replace("${company}", full_company_name)
                    if "${director}" in v:
                        v = v.replace("${director}", director)
                    if "${chief}" in v:
                        v = v.replace("${chief}", director)
                    if "${month}" in v:
                        v = v.replace("${month}", get_month_name_ukr(target_month))
                    if "${year}" in v:
                        v = v.replace("${year}", str(target_year))
                    if "${start_date}" in v:
                        v = v.replace("${start_date}", first_date_str)
                    if "${end_date}" in v:
                        v = v.replace("${end_date}", last_date_str)
                    cell.value = v
        
        company_df = df[df[col_company].apply(lambda x: str(x).strip() == company_str)]
        drivers = sorted(company_df[col_driver].unique())
        
        # Find initial razom_row
        razom_row = None
        for r in range(13, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "Разом":
                razom_row = r
                break
        
        if not razom_row:
            print(f"Error: 'Разом' row not found in template for {company_str}.")
            return

        while len(drivers) > (razom_row - 13) // 4:
            merged_ranges = list(ws.merged_cells.ranges)
            shifted_merges = []
            for m in merged_ranges:
                if m.min_row >= razom_row:
                    ws.unmerge_cells(str(m))
                    shifted_merges.append(m)
            
            ws.insert_rows(razom_row, 4)
            
            for m in shifted_merges:
                m.shift(0, 4)
                ws.merge_cells(str(m))
                
            for i in range(4):
                copy_row_style(ws, 13 + i, razom_row + i, 13, razom_row)
                
            for m in merged_ranges:
                if m.min_row >= 13 and m.max_row <= 16:
                    new_m = copy(m)
                    new_m.shift(0, razom_row - 13)
                    ws.merge_cells(str(new_m))
            
            razom_row += 4

        # Update SUM formulas in "Разом" row to reference the full driver range (row 13 to razom_row - 1)
        last_driver_row = razom_row - 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=razom_row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = re.sub(r'(?P<col>[A-Z]+)13:(?P=col)20', rf'\g<col>13:\g<col>{last_driver_row}', cell.value)
        
        current_row = 13
        
        for driver in drivers:
            short_name = format_driver_initials(driver)
            for r in range(current_row, current_row + 4):
                c = ws.cell(row=r, column=2)
                if isinstance(c.value, str) and "${driver}" in c.value:
                    c.value = c.value.replace("${driver}", short_name)
                    c.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal=c.alignment.horizontal, vertical=c.alignment.vertical)
                elif c.value == "${driver}":
                    c.value = short_name
                    c.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal=c.alignment.horizontal, vertical=c.alignment.vertical)
            
            driver_df = company_df[company_df[col_driver] == driver]
            trips = []
            for _, row_data in driver_df.iterrows():
                st = row_data[col_start]
                en = row_data[col_end]
                if pd.notna(st) and pd.notna(en):
                    try:
                        trips.append((pd.to_datetime(st).date(), pd.to_datetime(en).date()))
                    except:
                        pass
            
            # Find vacations for this driver
            driver_norm = normalize_name(driver)
            driver_vacations = [v for v in parsed_vacations if v["name"] == driver_norm and matches_company(v.get("company", ""), company_str)]
            
            for day in range(1, last_day + 1):
                current_date = datetime.date(target_year, target_month, day)
                
                # Check for matching vacation first
                matching_vacation = None
                for v in driver_vacations:
                    if v["start"] <= current_date <= v["end"]:
                        matching_vacation = v
                        break
                
                if matching_vacation:
                    v_type = matching_vacation["type"]
                    if v_type == "без збереження заробітної плати":
                        val = "НА"
                    elif v_type == "щорічна відпустка":
                        val = "В"
                    else:
                        val = "НА" if "без збереження" in v_type.lower() else "В"
                else:
                    is_engaged = False
                    for t_start, t_end in trips:
                        if t_start <= current_date <= t_end:
                            is_engaged = True
                            break
                    
                    is_weekday = current_date.weekday() < 5
                    
                    if is_engaged:
                        val = "ВД"
                    elif is_weekday:
                        val = "Р"
                    else:
                        val = "ВВ"
                
                if day <= 15:
                    col = day + 4
                    r = current_row
                else:
                    col = day - 15 + 4
                    r = current_row + 2
                    
                try:
                    ws.cell(row=r, column=col, value=val)
                except AttributeError:
                    pass
                
            current_row += 4
            
        for r in range(current_row, ws.max_row):
            c = ws.cell(row=r, column=2)
            if isinstance(c.value, str) and "${driver}" in c.value:
                for i in range(4):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=r+i, column=col, value="")
                break

        # Update day headers for days 29, 30, 31 in row 11
        for day in range(29, 32):
            col = day - 15 + 4
            if day <= last_day:
                ws.cell(row=11, column=col, value=day)
            else:
                ws.cell(row=11, column=col, value="X")

        out_name = f"Табель обліку робочого часу {company_str} {mmyy_str}.xlsx"
        out_path = os.path.join(output_dir, out_name)
        
        print(f"  Saving to {out_path} ...")
        try:
            wb.save(out_path)
            wb.close()
        except Exception as e:
            print(f"Error saving {out_path}: {e}")
            
    print("Done!")

if __name__ == "__main__":
    try:
        main()
    finally:
        try:
            input("\nНатисніть Enter для виходу / Press Enter to exit...")
        except (KeyboardInterrupt, EOFError):
            pass
