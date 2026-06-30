import os
import sys
import subprocess

# Auto-install missing dependencies (openpyxl and pywin32)
try:
    import openpyxl
except ImportError:
    print("Missing required dependency 'openpyxl'. Attempting to install automatically...", flush=True)
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        print("Successfully installed openpyxl!", flush=True)
    except Exception as e:
        print(f"Error installing openpyxl automatically: {e}", file=sys.stderr, flush=True)
        print("Please install it manually using: pip install openpyxl", file=sys.stderr, flush=True)
        sys.exit(1)

try:
    import win32com.client
except ImportError:
    print("Missing required dependency 'pywin32'. Attempting to install automatically...", flush=True)
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pywin32"])
        import win32com.client
        print("Successfully installed pywin32!", flush=True)
    except Exception as e:
        print(f"Error installing pywin32 automatically: {e}", file=sys.stderr, flush=True)
        print("Please install it manually using: pip install pywin32", file=sys.stderr, flush=True)
        sys.exit(1)

import glob
from pathlib import Path
from datetime import datetime
import argparse
import winsound

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)

# Find source file dynamically
source_files = glob.glob(os.path.join(ROOT_DIR, "*.xls*"))
source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
if len(source_files) == 1:
    SOURCE_FILE = source_files[0]
else:
    SOURCE_FILE = ""

def find_template_file(filename, fallback_dir="шаблони"):
    for path in Path(ROOT_DIR).rglob(filename):
        parts = [p.lower() for p in path.parts]
        if path.is_file() and not any(p.startswith('.') or p in ('venv', 'env', 'build', 'dist', 'node_modules', '__pycache__') for p in parts):
            return str(path)
    return os.path.join(ROOT_DIR, fallback_dir, filename)

TEMPLATE_FILE = find_template_file("шаблон-реєстру-зіа.xlsx")
TEMPLATE_FILE_ZETTRA = find_template_file("шаблон-реєстру-зет.xlsx")
REGISTRIES_BASE_PATH = os.path.join(ROOT_DIR, "реєстри")
os.makedirs(REGISTRIES_BASE_PATH, exist_ok=True)

# Define company names
COMPANY_AVTOTRANS = "ЗІАВТОТРАНС"
COMPANY_ZETTRA = "ЗЕТТРА"

# Define column mappings for replacements
COLUMN_MAPPINGS = {
    'customer': 'замовник',
    'companyshortname': 'замовник',
    'invoice': 'рахунок',
    'freight': 'вантаж (родовий відмінок)',
    'route': 'маршрут згідно товаро-транспортній накладній',
    'lastdate': 'розвантаження дата',
    'last date': 'розвантаження дата'
}

# Define table column mappings
TABLE_COLUMN_MAPPINGS = {
    'Дата розвантаження': 'розвантаження дата',
    'Дата навантаження': 'навантаження дата',
    '№ авто': 'авто',
    '№ причеп': 'причіп',
    'Вага навантаження, т': 'навантаження вага тонни',
    '№ ТТН': 'номер ттн',
    'Вага розвантаження, т': 'розвантаження вага тонни',
    'Ставка за 1 т, без ПДВ': 'ціна за тонну без ПДВ',
    'Тариф без ПДВ, грн': 'ціна за тонну без ПДВ',
    'ПІБ водія': 'водій',
    'Водій': 'водій',
    'Пункт навантаження': 'маршрут згідно товаро-транспортній накладній',
    'Пункт вивантаження': 'маршрут згідно товаро-транспортній накладній',
    'Пункт розвантаження': 'маршрут згідно товаро-транспортній накладній',
    'Культура': 'вантаж (родовий відмінок)'
}

TABLE_COLUMN_MAPPINGS_KG = TABLE_COLUMN_MAPPINGS

def excel_serial_to_date(serial_value):
    """Convert an Excel date serial number (e.g. 46096) or datetime object to a 'dd.mm.yyyy' string.
    Returns the original value unchanged if it cannot be converted."""
    from datetime import datetime, timedelta
    try:
        if serial_value is None or str(serial_value).strip() == '':
            return serial_value
        if isinstance(serial_value, datetime):
            return serial_value.strftime('%d.%m.%Y')
        if isinstance(serial_value, str):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                try:
                    return datetime.strptime(serial_value, fmt).strftime('%d.%m.%Y')
                except ValueError:
                    continue
        serial = int(float(str(serial_value)))
        if serial < 1:  # Not a valid date serial
            return serial_value
        # Excel epoch is 30-Dec-1899; day 1 = 1-Jan-1900
        base = datetime(1899, 12, 30)
        date = base + timedelta(days=serial)
        return date.strftime('%d.%m.%Y')
    except (ValueError, TypeError, OverflowError):
        return serial_value

def convert_weight_to_kg(weight_str):
    """Convert weight from tons to kg (multiply by 1000)"""
    try:
        if weight_str and str(weight_str).strip():
            weight = float(str(weight_str).replace(',', '.'))
            return int(weight * 1000)
        return ''
    except (ValueError, AttributeError):
        return ''

def calculate_rate_with_vat(freight_str):
    """Calculate rate with VAT from freight (freight/100)"""
    try:
        if freight_str and str(freight_str).strip():
            freight = float(str(freight_str).replace(',', '.'))
            return round(freight / 1000, 2)
        return ''
    except (ValueError, AttributeError):
        return ''

def parse_route_information(route_str):
    """Parse route string to extract loading and unloading points"""
    if not route_str or str(route_str).strip() == '':
        return {'loading_point': '', 'unloading_point': ''}
    
    route = str(route_str).strip()
    
    # Look for common route separators: em dash (–), en dash (–), or regular hyphen (-)
    # Routes often use the Ukrainian typographic em dash: " – "
    separator = None
    for sep in [' \u2013 ', ' \u2014 ', ' - ']:
        if sep in route:
            separator = sep
            break

    if separator:
        parts = route.split(separator, 1)  # Split only on the first occurrence
        loading_point = parts[0].strip()
        unloading_point = parts[1].strip() if len(parts) > 1 else ''
        
        print(f"Parsed route: Loading='{loading_point}', Unloading='{unloading_point}'")
        return {
            'loading_point': loading_point,
            'unloading_point': unloading_point
        }
    else:
        # If no separator found, consider the whole string as loading point
        print(f"No route separator found, using entire string as loading point: '{route}'")
        return {
            'loading_point': route,
            'unloading_point': ''
        }

def convert_crop_to_nominative(crop_str):
    """Convert crop name from genitive case to nominative case"""
    if not crop_str or str(crop_str).strip() == '':
        return ''
    
    crop = str(crop_str).strip().lower()
    
    # Dictionary of genitive to nominative crop name conversions
    crop_conversions = {
        'ячміню': 'ячмінь',
        'пшениці': 'пшениця',
        'кукурудзи': 'кукурудза',
        'соняшнику': 'соняшник',
        'ріпаку': 'ріпак',
        'сої': 'соя',
        'овса': 'овес',
        'жита': 'жито',
        'гречки': 'гречка',
        'проса': 'просо',
        'льону': 'льон',
        'гороху': 'горох',
        'квасолі': 'квасоля',
        'чечевиці': 'чечевиця',
        'нуту': 'нут',
        'шроту':'шрот',
        'шроту соєвого':'шрот соєвий',
        'шроту соняшникового':'шрот соняшниковий'
    }
    
    # Check if we have a direct conversion
    if crop in crop_conversions:
        result = crop_conversions[crop]
        print(f"Converted crop name: '{crop}' -> '{result}'")
        return result
    
    # If no conversion found, return original value with proper capitalization
    result = crop.capitalize()
    print(f"No conversion found for crop '{crop}', using: '{result}'")
    return result

def create_company_short_name(full_company_name):
    """Create short company name by keeping innermost quoted text and abbreviating everything else"""
    if not full_company_name or str(full_company_name).strip() == '':
        print("[SHORT_NAME] Empty input, returning empty string")
        return ""
    
    text = str(full_company_name).strip()
    print(f"\n[SHORT_NAME] Input: '{text}'")
    print(f"[SHORT_NAME] Input length: {len(text)}")
    
    # Debug: Show all quote characters and their positions
    print("[SHORT_NAME] Quote character positions:")
    for i, char in enumerate(text):
        if char in ['"', '«', '»']:
            print(f"  Position {i}: '{char}' (ord={ord(char)})")
    
    import re
    
    # Find all quote positions
    quote_positions = [i for i, c in enumerate(text) if c == '"']
    print(f"[SHORT_NAME] Found {len(quote_positions)} quote characters at positions: {quote_positions}")
    
    # Find the innermost (deepest nested) quoted text
    # The innermost quote pair is the last opening quote and its matching closing quote
    innermost_text = None
    innermost_start = -1
    innermost_end = -1
    
    if len(quote_positions) >= 2:
        # Find the last pair - look for the last opening quote that has a closing quote after it
        for i in range(len(quote_positions) - 2, -1, -1):
            # This could be an opening quote if there's at least one more quote after it
            if i + 1 < len(quote_positions):
                innermost_start = quote_positions[i]
                innermost_end = quote_positions[i + 1]
                innermost_text = text[innermost_start + 1:innermost_end]
                print(f"[SHORT_NAME] Innermost quoted text: '{innermost_text}' (positions {innermost_start}-{innermost_end})")
                break
    
    if innermost_text is None:
        print("[SHORT_NAME] No quotes found or unable to parse, returning original text")
        return text
    
    # Get everything before the innermost quoted text (including the opening quote)
    text_before_innermost = text[:innermost_start].strip()
    print(f"[SHORT_NAME] Text before innermost: '{text_before_innermost}'")
    
    # Remove all quote characters from text_before_innermost
    text_to_abbreviate = re.sub(r'[«»"]', '', text_before_innermost).strip()
    print(f"[SHORT_NAME] Text to abbreviate (quotes removed): '{text_to_abbreviate}'")
    
    if not text_to_abbreviate:
        print("[SHORT_NAME] Nothing to abbreviate, returning innermost text only")
        return f'"{innermost_text}"'
    
    # Split into parts and abbreviate
    parts = text_to_abbreviate.split()
    print(f"[SHORT_NAME] Parts to process: {parts}")
    abbreviated_parts = []
    
    i = 0
    while i < len(parts):
        # Check if this is a known multi-word entity type
        remaining_text = ' '.join(parts[i:])
        
        if remaining_text.startswith("ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ"):
            abbreviated_parts.append("ТОВ")
            print(f"[SHORT_NAME] Found 'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ' -> 'ТОВ'")
            i += 4
        elif remaining_text.startswith("ПРИВАТНЕ ПІДПРИЄМСТВО"):
            abbreviated_parts.append("ПП")
            print(f"[SHORT_NAME] Found 'ПРИВАТНЕ ПІДПРИЄМСТВО' -> 'ПП'")
            i += 2
        elif remaining_text.startswith("ФІЗИЧНА ОСОБА - ПІДПРИЄМЕЦЬ"):
            abbreviated_parts.append("ФОП")
            print(f"[SHORT_NAME] Found 'ФІЗИЧНА ОСОБА - ПІДПРИЄМЕЦЬ' -> 'ФОП'")
            i += 4
        elif remaining_text.startswith("ФІЗИЧНА ОСОБА ПІДПРИЄМЕЦЬ"):
            abbreviated_parts.append("ФОП")
            print(f"[SHORT_NAME] Found 'ФІЗИЧНА ОСОБА ПІДПРИЄМЕЦЬ' -> 'ФОП'")
            i += 3
        elif remaining_text.startswith("ДЕРЖАВНЕ ПІДПРИЄМСТВО"):
            abbreviated_parts.append("ДП")
            print(f"[SHORT_NAME] Found 'ДЕРЖАВНЕ ПІДПРИЄМСТВО' -> 'ДП'")
            i += 2
        elif remaining_text.startswith("КОМУНАЛЬНЕ ПІДПРИЄМСТВО"):
            abbreviated_parts.append("КП")
            print(f"[SHORT_NAME] Found 'КОМУНАЛЬНЕ ПІДПРИЄМСТВО' -> 'КП'")
            i += 2
        elif remaining_text.startswith("ІНОЗЕМНЕ ПІДПРИЄМСТВО"):
            abbreviated_parts.append("ІП")
            print(f"[SHORT_NAME] Found 'ІНОЗЕМНЕ ПІДПРИЄМСТВО' -> 'ІП'")
            i += 2
        elif remaining_text.startswith("ПІДПРИЄМСТВО З ІНОЗЕМНИМИ ІНВЕСТИЦІЯМИ"):
            abbreviated_parts.append("ПІІ")
            print(f"[SHORT_NAME] Found 'ПІДПРИЄМСТВО З ІНОЗЕМНИМИ ІНВЕСТИЦІЯМИ' -> 'ПІІ'")
            i += 4
        elif remaining_text.startswith("НЕПРИБУТКОВА ОРГАНІЗАЦІЯ"):
            abbreviated_parts.append("НПО")
            print(f"[SHORT_NAME] Found 'НЕПРИБУТКОВА ОРГАНІЗАЦІЯ' -> 'НПО'")
            i += 2
        elif remaining_text.startswith("АКЦІОНЕРНЕ ТОВАРИСТВО"):
            abbreviated_parts.append("АТ")
            print(f"[SHORT_NAME] Found 'АКЦІОНЕРНЕ ТОВАРИСТВО' -> 'АТ'")
            i += 2
        elif remaining_text.startswith("ПУБЛІЧНЕ АКЦІОНЕРНЕ ТОВАРИСТВО"):
            abbreviated_parts.append("ПАТ")
            print(f"[SHORT_NAME] Found 'ПУБЛІЧНЕ АКЦІОНЕРНЕ ТОВАРИСТВО' -> 'ПАТ'")
            i += 3
        elif remaining_text.startswith("ПРИВАТНЕ АКЦІОНЕРНЕ ТОВАРИСТВО"):
            abbreviated_parts.append("ПрАТ")
            print(f"[SHORT_NAME] Found 'ПРИВАТНЕ АКЦІОНЕРНЕ ТОВАРИСТВО' -> 'ПрАТ'")
            i += 3
        elif remaining_text.startswith("ВІДКРИТЕ АКЦІОНЕРНЕ ТОВАРИСТВО"):
            abbreviated_parts.append("ВАТ")
            print(f"[SHORT_NAME] Found 'ВІДКРИТЕ АКЦІОНЕРНЕ ТОВАРИСТВО' -> 'ВАТ'")
            i += 3
        elif remaining_text.startswith("ЗАКРИТЕ АКЦІОНЕРНЕ ТОВАРИСТВО"):
            abbreviated_parts.append("ЗАТ")
            print(f"[SHORT_NAME] Found 'ЗАКРИТЕ АКЦІОНЕРНЕ ТОВАРИСТВО' -> 'ЗАТ'")
            i += 3
        elif remaining_text.startswith("ПОВНЕ ТОВАРИСТВО"):
            abbreviated_parts.append("ПТ")
            print(f"[SHORT_NAME] Found 'ПОВНЕ ТОВАРИСТВО' -> 'ПТ'")
            i += 2
        elif remaining_text.startswith("СІМЕЙНЕ ФЕРМЕРСЬКЕ ГОСПОДАРСТВО"):
            abbreviated_parts.append("СФГ")
            print(f"[SHORT_NAME] Found 'СІМЕЙНЕ ФЕРМЕРСЬКЕ ГОСПОДАРСТВО' -> 'СФГ'")
            i += 3
        elif remaining_text.startswith("ОБ'ЄДНАННЯ СПІВВЛАСНИКІВ БАГАТОКВАРТИРНОГО БУДИНКУ"):
            abbreviated_parts.append("ОСББ")
            print(f"[SHORT_NAME] Found 'ОБ'ЄДНАННЯ СПІВВЛАСНИКІВ БАГАТОКВАРТИРНОГО БУДИНКУ' -> 'ОСББ'")
            i += 4
        elif remaining_text.startswith("ПІДПРИЄМСТВО ОБ'ЄДНАННЯ ГРОМАДЯН"):
            abbreviated_parts.append("ПОГ")
            print(f"[SHORT_NAME] Found 'ПІДПРИЄМСТВО ОБ'ЄДНАННЯ ГРОМАДЯН' -> 'ПОГ'")
            i += 3
        elif remaining_text.startswith("ТОРГОВИЙ ДІМ"):
            abbreviated_parts.append("ТД")
            print(f"[SHORT_NAME] Found 'ТОРГОВИЙ ДІМ' -> 'ТД'")
            i += 2
        elif remaining_text.startswith("КОНДИТЕРСЬКА ФАБРРИКА"):
            abbreviated_parts.append("КФ")
            print(f"[SHORT_NAME] Found 'КОНДИТЕРСЬКА ФАБРРИКА' -> 'КФ'")
            i += 2
        else:
            # For other words, take first letter
            word = parts[i]
            if word and word != '-':
                abbreviated_parts.append(word[0].upper())
                print(f"[SHORT_NAME] Abbreviating '{word}' -> '{word[0].upper()}'")
            i += 1
    
    # Join abbreviations
    abbreviation = ' '.join(abbreviated_parts)
    print(f"[SHORT_NAME] Final abbreviation: '{abbreviation}'")
    
    # Now reconstruct with proper quote structure
    # Count how many quotes were before the innermost
    quotes_before = text[:innermost_start].count('"')
    print(f"[SHORT_NAME] Found {quotes_before} quotes before innermost")
    
    # Build result with proper nesting
    if quotes_before == 0:
        # Simple case: ABB "innermost"
        result = f'{abbreviation} "{innermost_text}"'
    elif quotes_before == 1:
        # One level of nesting: ABB "last_part "innermost"
        # Split abbreviation - last part should be inside quotes with innermost
        abb_parts = abbreviated_parts
        if len(abb_parts) > 1:
            # Put last abbreviation inside quotes with innermost
            outer_parts = ' '.join(abb_parts[:-1])
            inner_part = abb_parts[-1]
            result = f'{outer_parts} "{inner_part} "{innermost_text}"'
        else:
            # Only one part, put it inside quotes
            result = f'"{abbreviation} "{innermost_text}"'
    else:
        # Multiple levels - add opening quotes and closing quotes
        # For simplicity, put last part inside all nested quotes
        abb_parts = abbreviated_parts
        if len(abb_parts) > quotes_before:
            outer_parts = ' '.join(abb_parts[:-quotes_before])
            inner_parts = ' '.join(abb_parts[-quotes_before:])
            result = outer_parts + ' "' * quotes_before + inner_parts + ' "' + innermost_text + '"' * quotes_before
        else:
            result = '"' * quotes_before + abbreviation + ' "' + innermost_text + '"' * quotes_before
    
    print(f"[SHORT_NAME] Final result: '{result}'")
    return result

def get_full_customer_name(workbook, short_name):
    return short_name

def get_customer_template_files(workbook, customer_name):
    """Look up customer-specific template files in the Контрагенти sheet"""
    try:
        if "Контрагенти" not in workbook.sheetnames:
            return None
        customers_sheet = workbook["Контрагенти"]
        # Find the customers table
        table = None
        for t in customers_sheet.tables.values():
            if t.name == "customers":
                table = t
                break
        if not table:
            print("Warning: 'customers' table not found in Контрагенти sheet")
            return None
            
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [customers_sheet.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
        headers = [str(h).strip().lower() if h is not None else "" for h in headers]
        
        try:
            contractor_col = headers.index("контрагент")
        except ValueError:
            print("Warning: 'контрагент' column not found in customers table")
            return None
            
        try:
            template_files_col = headers.index("файл реєстру (зет;зіа)")
        except ValueError:
            print("Info: 'файл реєстру (зет;зіа)' column not found in customers table")
            return None
            
        for row in range(min_row + 1, max_row + 1):
            cell_val = customers_sheet.cell(row=row, column=min_col + contractor_col).value
            if cell_val and str(cell_val).strip() == customer_name.strip():
                template_files = customers_sheet.cell(row=row, column=min_col + template_files_col).value
                if template_files:
                    template_files_str = str(template_files).strip()
                    print(f"Found template files: {template_files_str}")
                    files = [f.strip() for f in template_files_str.split(';') if f.strip()]
                    return files
                else:
                    print(f"No template files specified for customer '{customer_name}'")
                    return None
                    
        print(f"Customer '{customer_name}' not found in Контрагенти sheet")
        return None
    except Exception as e:
        print(f"Warning: Error looking up customer template files: {str(e)}")
        return None

def get_customer_edrpou(workbook, customer_name):
    return ""

def replace_template_text(workbook, search_text, replace_text):
    """Replace specific words in all sheets of the template while preserving other text"""
    try:
        # Special handling for invoice field when empty
        if search_text == 'invoice' and not replace_text:
            replace_text = "______"
            
        if not replace_text and search_text != 'invoice':  # Skip empty replacements except for invoice
            print(f"Warning: Empty replacement value for '{search_text}', skipping...")
            return
            
        for sheet in workbook.Worksheets:
            # Find cells containing the search text using LookAt=2 (xlPart) to ensure substring match
            found = sheet.Cells.Find(search_text, LookAt=2)
            first_address = None
            
            while found and (first_address is None or found.Address != first_address):
                if first_address is None:
                    first_address = found.Address
                
                # Get the current cell text
                current_text = found.Text
                
                # Replace only the specific word while preserving other text
                new_text = current_text.replace(search_text, str(replace_text))
                
                # Update the cell only if text changed
                if new_text != current_text:
                    found.Value = new_text
                    print(f"Updated cell {found.Address}: '{current_text}' -> '{new_text}'")
                
                # Find next occurrence
                found = sheet.Cells.FindNext(found)
                
                # Break if we've wrapped around to the beginning
                if not found or (first_address and found.Address == first_address):
                    break
                    
        print(f"Replaced all occurrences of '{search_text}' with '{replace_text}'")
    except Exception as e:
        print(f"Warning: Error replacing text: {str(e)}")

def replace_template_text_kg_mode(workbook):
    """Replace template text for kg mode"""
    try:
        replacements = {
            'Вага навантаження, т': 'Вага навантаження, кг',
            'Вага розвантаження, т': 'Вага розвантаження, кг',
            'Ставка за 1 т, без ПДВ': 'Ставка за 1 кг, з ПДВ',
            'Сумарна вага, т': 'Сумарна вага, кг',
            'Загалом сума без ПДВ': 'Загалом сума з ПДВ'
        }
        
        for search_text, replace_text in replacements.items():
            replace_template_text(workbook, search_text, replace_text)
            
        print("Completed kg mode text replacements")
    except Exception as e:
        print(f"Warning: Error in kg mode text replacement: {str(e)}")

def remove_vat_total_row(sheet):
    """Clear the content and borders of the row containing 'Всього сума з ПДВ:' cell"""
    try:
        found = sheet.Cells.Find('Всього сума з ПДВ:')
        if found:
            row_to_clear = found.Row
            # Clear content
            sheet.Rows(row_to_clear).ClearContents()
            # Clear borders
            sheet.Rows(row_to_clear).Borders.LineStyle = 0  # 0 = xlNone
            print(f"Cleared content and borders of row {row_to_clear} containing 'Всього сума з ПДВ:'")
            return True
        else:
            print("Warning: Could not find 'Всього сума з ПДВ:' cell to clear")
            return False
    except Exception as e:
        print(f"Warning: Error clearing VAT total row: {str(e)}")
        return False

def get_save_path(sheet, table, client_name, latest_date, filtered_gid=None, invoice_number=None, company=None, fileformat='xlsx'):
    if not invoice_number:
        invoice_number = "NO_INVOICE"
    
    # Determine the company code: ЗІА or ЗЕТ
    company_upper = str(company).upper() if company else ""
    if 'ЗІА' in company_upper:
        company_code = "ЗІА"
    elif 'ЗЕТ' in company_upper:
        company_code = "ЗЕТ"
    else:
        # Fallback to carrier names if not directly containing ZIA/ZET
        if 'ЗІАВТОТРАНС' in company_upper:
            company_code = "ЗІА"
        elif 'ЗЕТТРА' in company_upper:
            company_code = "ЗЕТ"
        else:
            company_code = str(company)
            
    # Clean client name and invoice number for Windows filename compatibility
    clean_client = str(client_name)
    clean_invoice = str(invoice_number)
    
    # Remove characters invalid in Windows filenames: \ / : * ? " < > |
    invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    for char in invalid_chars:
        clean_client = clean_client.replace(char, '')
        clean_invoice = clean_invoice.replace(char, '_')
        company_code = company_code.replace(char, '_')
    clean_client = clean_client.strip()
    clean_invoice = clean_invoice.strip()
    company_code = company_code.strip()

    date_suffix = ""
    if latest_date:
        if hasattr(latest_date, 'strftime'):
            date_suffix = f"-{latest_date.strftime('%d%m%y')}"
        elif isinstance(latest_date, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    dt = datetime.strptime(latest_date, fmt)
                    date_suffix = f"-{dt.strftime('%d%m%y')}"
                    break
                except ValueError:
                    continue
    
    filename = f"{clean_client}-{company_code}-{clean_invoice}{date_suffix}.{fileformat}"
    
    os.makedirs(REGISTRIES_BASE_PATH, exist_ok=True)
    gid_path = os.path.join(REGISTRIES_BASE_PATH, filename)
    return {'gid_path': gid_path}

def find_table_range(sheet):
    """Find the table range in template based on 'Дата розвантаження' marker"""
    try:
        # Find the cell containing 'Дата розвантаження'
        found = sheet.Cells.Find('Дата розвантаження')
        if not found:
            print("ERROR: Could not find 'Дата розвантаження' marker in template")
            return None
            
        start_row = found.Row
        print(f"Found table start marker at row {start_row}")
        
        # The table has 3 rows by default (header + two data rows)
        template_rows = 3
        
        return {
            'start_row': start_row,
            'template_rows': template_rows
        }
    except Exception as e:
        print(f"ERROR finding table range: {str(e)}")
        return None

def adjust_template_rows(sheet, table_info, required_rows, is_kg_mode=False):
    """Adjust the number of rows in template to match required rows"""
    try:
        start_row = table_info['start_row']
        template_rows = table_info['template_rows']
        current_data_rows = template_rows - 1  # Exclude header row
        
        print(f"\nAdjusting template rows: Current={current_data_rows}, Required={required_rows}")
        
        # Find the sum formula cell (use appropriate header based on mode)
        sum_header = 'Сумарна вага, кг' if is_kg_mode else 'Сумарна вага, т'
        header_cell = sheet.Cells.Find(sum_header)
        has_sum_header = header_cell is not None
        if not has_sum_header:
            print(f"Info: '{sum_header}' cell not found - this template might not have sum formulas")
            
        # Get the column letter for weight values
        weight_col = None
        header_row = sheet.Rows(start_row)
        weight_header = 'Вага розвантаження, кг' if is_kg_mode else 'Вага розвантаження, т'
        weight_cell = header_row.Find(weight_header)
        if weight_cell:
            weight_col = get_column_letter(weight_cell.Column)
        else:
            print(f"Info: Could not find '{weight_header}' column - template might use different column names")

        if required_rows > current_data_rows:
            # Need to add rows
            rows_to_add = required_rows - current_data_rows
            print(f"Adding {rows_to_add} rows")
            
            # Insert rows from bottom to top to avoid shifting issues
            for i in range(rows_to_add):
                # Copy the first data row and insert at the end of current data
                source_range = sheet.Rows(start_row + 1)  # Copy the first data row
                target_row = start_row + current_data_rows + 1  # Always insert at the end
                source_range.Copy()
                sheet.Rows(target_row).Insert()
                print(f"Inserted row at position {target_row}")
                
        elif required_rows < current_data_rows:
            # Need to remove rows
            rows_to_remove = current_data_rows - required_rows
            print(f"Removing {rows_to_remove} rows")
            
            for i in range(rows_to_remove):
                # Delete rows from bottom up
                sheet.Rows(start_row + required_rows + 1).Delete()

        # After all row operations, find and update ALL SUM formulas in the table area
        # Use regex so we catch every column's totals (weight, price, etc.) regardless of template layout
        import re
        first_data_row = start_row + 1
        last_data_row = start_row + required_rows
        search_start_row = start_row
        search_end_row = last_data_row + 10  # Search a bit beyond the table
        sum_formulas_updated = 0

        print(f"Looking for SUM formulas to update in rows {search_start_row}-{search_end_row}...")
        for search_row in range(search_start_row, search_end_row + 1):
            for search_col in range(1, 30):  # Search first 30 columns
                cell = sheet.Cells(search_row, search_col)
                formula = str(cell.Formula)
                if 'SUM(' not in formula:
                    continue
                # Match patterns like SUM(K6:K7) or SUM(M5:M8) etc.
                # Groups: (1)=open col, (2)=open row, (3)=close col, (4)=close row
                def replace_sum_range(m):
                    col_letter = m.group(1)
                    row1 = int(m.group(2))
                    row2 = int(m.group(4))
                    # Only update if the range overlaps the original data area
                    orig_first = start_row + 1
                    orig_last = start_row + current_data_rows
                    if row1 >= orig_first and row2 <= orig_last + 10:
                        return f'SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
                    return m.group(0)  # Leave unchanged
                new_formula = re.sub(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', replace_sum_range, formula)
                if new_formula != formula:
                    old_formula = formula
                    cell.Formula = new_formula
                    print(f"Updated SUM formula at {cell.Address}")
                    print(f"  Old: {old_formula}")
                    print(f"  New: {new_formula}")
                    sum_formulas_updated += 1

        if sum_formulas_updated == 0:
            print("Info: No SUM formulas found to update in this template")
        else:
            print(f"Updated {sum_formulas_updated} SUM formula(s)")
        
        # Also handle sumloadingweight placeholder - always uses column E
        print("Looking for 'sumloadingweight' placeholder to replace with loading weight sum...")
        sumloadingweight_found = False
        
        # Search for sumloadingweight placeholder in the entire sheet
        found = sheet.Cells.Find('sumloadingweight')
        if found:
            first_data_row = start_row + 1
            last_data_row = start_row + required_rows
            loading_weight_formula = f'=ROUND(SUM(F{first_data_row}:E{last_data_row}), 2)'
            found.Value = loading_weight_formula
            print(f"Replaced 'sumloadingweight' at {found.Address} with: {loading_weight_formula}")
            sumloadingweight_found = True
        
        if not sumloadingweight_found:
            print("Note: No 'sumloadingweight' placeholder found in template")
                
        print(f"Template now has {required_rows} data rows")
        return True
    except Exception as e:
        print(f"ERROR adjusting template rows: {str(e)}")
        return False

def get_column_letter(col_num):
    """Convert column number to letter (1 = A, 2 = B, etc.)"""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def hide_loading_weight_column(sheet, table_info, is_kg_mode=False):
    """Hide the loading weight column in the template"""
    try:
        start_row = table_info['start_row']
        column_to_hide = 'Вага навантаження, кг' if is_kg_mode else 'Вага навантаження, т'
        
        # Find the column
        found = sheet.Rows(start_row).Find(column_to_hide)
        if found:
            col_idx = found.Column
            # Hide the entire column
            sheet.Columns(col_idx).Hidden = True
            print(f"Hidden column '{column_to_hide}' at position {col_idx}")
            return True
        else:
            print(f"Warning: Could not find column '{column_to_hide}' to hide")
            return False
    except Exception as e:
        print(f"ERROR hiding loading weight column: {str(e)}")
        return False

def fill_table_data(sheet, table_info, source_data, source_columns, is_kg_mode=False, skip_loading_weight=False):
    """Fill the template table with data from source"""
    try:
        start_row = table_info['start_row']
        
        # Choose the appropriate column mappings
        column_mappings = TABLE_COLUMN_MAPPINGS_KG if is_kg_mode else TABLE_COLUMN_MAPPINGS
        
        # If skip_loading_weight is True, create a copy without the loading weight column
        if skip_loading_weight:
            column_mappings = column_mappings.copy()
            loading_weight_key = 'Вага навантаження, кг' if is_kg_mode else 'Вага навантаження, т'
            if loading_weight_key in column_mappings:
                del column_mappings[loading_weight_key]
                print(f"Excluding '{loading_weight_key}' from data filling")
        
        print(f"\nFilling table data (kg mode: {is_kg_mode}):")
        # For each row of data
        for row_idx, row_data in enumerate(source_data):
            target_row = start_row + row_idx + 1  # +1 to skip header
            
            # Fill each column
            for template_col, source_col in column_mappings.items():
                if template_col in ('ПІБ водія', 'Водій'):
                    continue
                if source_col in source_columns or source_col in row_data:
                    value = row_data.get(source_col, '')
                    
                    # Special handling for route parsing
                    if template_col == 'Пункт навантаження':
                        # Parse loading point from route
                        route_str = row_data.get('маршрут згідно товаро-транспортній накладній', '')
                        route_info = parse_route_information(route_str)
                        value = route_info['loading_point']
                    elif template_col == 'Пункт вивантаження' or template_col == 'Пункт розвантаження':
                        # Parse unloading point from route (both names refer to the same thing)
                        route_str = row_data.get('маршрут згідно товаро-транспортній накладній', '')
                        route_info = parse_route_information(route_str)
                        value = route_info['unloading_point']
                    elif template_col == 'Культура':
                        # Convert crop name from genitive to nominative case
                        freight_str = row_data.get('вантаж', '')
                        if not freight_str:
                            freight_str = row_data.get('вантаж (родовий відмінок)', '')
                        value = convert_crop_to_nominative(freight_str)
                    elif template_col == 'ПІБ водія' or template_col == 'Водій':
                        # Format driver name from full name to initials
                        driver_str = row_data.get('водій', '')
                        value = format_driver_name(driver_str)
                    elif template_col in ('Дата розвантаження', 'Дата навантаження'):
                        # Convert Excel date serial number to formatted date string
                        value = excel_serial_to_date(value)
                    elif template_col in ('Тариф без ПДВ, грн', 'Ставка за 1 т, без ПДВ', 'Ставка за 1 кг, з ПДВ'):
                        # Round price/rate to 2 decimal places
                        try:
                            value = round(float(value), 2)
                        except (ValueError, TypeError):
                            pass  # Keep original value if conversion fails
                    
                    # Apply conversions for kg mode
                    if is_kg_mode:
                        if template_col == 'Вага навантаження, кг':
                            # Convert from tons to kg
                            value = convert_weight_to_kg(value)
                        elif template_col == 'Вага розвантаження, кг':
                            # Convert from tons to kg
                            value = convert_weight_to_kg(value)
                        elif template_col == 'Ставка за 1 кг, з ПДВ':
                            # Calculate rate with VAT from freight
                            value = calculate_rate_with_vat(value)
                    
                    # Find the column in the template
                    found = sheet.Rows(start_row).Find(template_col)
                    if found:
                        col_idx = found.Column
                        # Set the value
                        sheet.Cells(target_row, col_idx).Value = value
                        print(f"Set {template_col} at row {target_row} to '{value}'")
                    else:
                        print(f"Warning: Could not find column '{template_col}' in template")
                else:
                    print(f"Warning: Source column '{source_col}' not found in data")
                    
        return True
    except Exception as e:
        print(f"ERROR filling table data: {str(e)}")
        return False

def get_visible_rows_data(invoice_number=None):
    global SOURCE_FILE
    try:
        print("\n=== Starting to process Excel file ===")
        if not invoice_number:
            print("ERROR: Invoice number is required when manual selection is disabled.")
            return None
            
        print(f"Filtering by invoice number: {invoice_number}")
        
        # Verify SOURCE_FILE exists
        if not SOURCE_FILE or not os.path.exists(SOURCE_FILE):
            # Try to find *.xls* dynamically
            source_files = glob.glob(os.path.join(ROOT_DIR, "*.xls*"))
            source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
            if len(source_files) == 1:
                SOURCE_FILE = source_files[0]
            else:
                print(f"ERROR: Source file not found or multiple source files found in {ROOT_DIR}")
                sys.exit(1)
                
        print(f"Opening source file with openpyxl: {SOURCE_FILE}")
        wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
        
        if "Дебетовий" not in wb.sheetnames:
            print(f"ERROR: Could not find 'Дебетовий' sheet in workbook")
            return None
            
        sheet = wb["Дебетовий"]
        table = None
        for t in sheet.tables.values():
            if t.name == "Table1":
                table = t
                break
                
        if not table:
            print(f"ERROR: Could not find 'Table1' table in 'Дебетовий' sheet")
            return None
            
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
        headers = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(headers)]
        
        # Build columns mapping from name to column list index
        columns = {}
        company_col = None
        
        # Track all needed columns
        all_needed_columns = set(COLUMN_MAPPINGS.values()) | set(TABLE_COLUMN_MAPPINGS.values())
        # Ensure we also include 'товариство', 'вантаж', 'gid' if they exist
        all_needed_columns.add('товариство')
        all_needed_columns.add('вантаж')
        all_needed_columns.add('gid')
        
        for idx, col_name in enumerate(headers):
            if col_name == "товариство":
                company_col = idx
                print(f"Found 'товариство' column at position {idx + 1}")
            if col_name in all_needed_columns:
                columns[col_name] = idx
                print(f"Found '{col_name}' column at position {idx + 1}")
                
        if company_col is None:
            print("ERROR: Could not find 'товариство' column for template selection!")
            return None
            
        # Check for basic required columns
        basic_required = set(COLUMN_MAPPINGS.values()) | set(TABLE_COLUMN_MAPPINGS.values())
        missing_basic = [col for col in basic_required if col not in columns]
        if missing_basic:
            print(f"ERROR: Missing required columns: {', '.join(missing_basic)}")
            return None
            
        # Find the invoice column index
        try:
            invoice_col_idx = headers.index('рахунок')
        except ValueError:
            print("ERROR: Could not find 'рахунок' column for invoice filtering")
            return None
            
        # Filter rows by invoice number
        filtered_rows = []
        for r in range(min_row + 1, max_row + 1):
            invoice_val = sheet.cell(row=r, column=min_col + invoice_col_idx).value
            if invoice_val is not None and str(invoice_val).strip() == str(invoice_number).strip():
                filtered_rows.append(r)
                
        print(f"Found {len(filtered_rows)} rows matching invoice {invoice_number}")
        if not filtered_rows:
            print(f"ERROR: No rows found for invoice number {invoice_number}")
            return None
            
        print("\nCollecting data from matching rows:")
        companies_data = []
        latest_date = None
        customer_name_cache = {}
        
        for row_num in filtered_rows:
            row_data = {'row': row_num}
            
            # Get company for template selection
            company_val = sheet.cell(row=row_num, column=min_col + company_col).value
            company = str(company_val).strip() if company_val is not None else ''
            row_data['company'] = company
            
            # Collect all column values
            for col_name, col_idx in columns.items():
                val = sheet.cell(row=row_num, column=min_col + col_idx).value
                
                # Convert to string, handling None values and cleaning up float representations
                if val is None:
                    value = ''
                elif isinstance(val, (int, float)):
                    if isinstance(val, float) and val.is_integer():
                        value = str(int(val))
                    else:
                        value = str(val)
                elif isinstance(val, datetime):
                    value = val
                else:
                    value = str(val)
                    
                row_data[col_name] = value
                
                # Special handling for dates
                if col_name in ['розвантаження', 'розвантаження дата'] and val:
                    try:
                        parsed_date = None
                        if isinstance(val, datetime):
                            parsed_date = val
                        elif isinstance(val, (int, float)):
                            from datetime import timedelta
                            base = datetime(1899, 12, 30)
                            parsed_date = base + timedelta(days=int(val))
                        else:
                            str_val = str(val).strip()
                            try:
                                serial = int(float(str_val))
                                if serial >= 1:
                                    from datetime import timedelta
                                    base = datetime(1899, 12, 30)
                                    parsed_date = base + timedelta(days=serial)
                            except (ValueError, TypeError):
                                try:
                                    parsed_date = datetime.strptime(str_val, "%d.%m.%Y")
                                except ValueError:
                                    pass
                        if parsed_date:
                            if latest_date is None or parsed_date > latest_date:
                                latest_date = parsed_date
                    except Exception as e:
                        print(f"Warning: Could not process date in row {row_num}: {str(e)}")
                        
            # Get full customer name
            if 'замовник' in row_data:
                customer_short_name = row_data['замовник']
                if customer_short_name not in customer_name_cache:
                    full_client_name = get_full_customer_name(wb, customer_short_name)
                    customer_name_cache[customer_short_name] = full_client_name
                else:
                    full_client_name = customer_name_cache[customer_short_name]
                row_data['full_customer_name'] = full_client_name
                
            companies_data.append(row_data)
            print(f"Row {row_num}: Company '{company}', Processing {len(row_data)} fields")
            
        print(f"\nSuccessfully collected data from {len(companies_data)} rows")
        if latest_date:
            print(f"Latest unload date found: {latest_date.strftime('%d-%m-%Y')}")
            
        filtered_gid = companies_data[0].get('gid') if companies_data else None
        
        return companies_data, None, wb, sheet, latest_date, columns, filtered_gid
        
    except Exception as e:
        print(f"\nERROR in get_visible_rows_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def get_template_path(company, customer_name=None, source_workbook=None):
    """Get template path for the given company, with customer-specific override if available"""
    print(f"\nDetermining template for company: {company}")
    
    company_upper = str(company).upper()
    is_avtotrans = 'ЗІА' in company_upper or company_upper == COMPANY_AVTOTRANS.upper()
    is_zettra = 'ЗЕТ' in company_upper or company_upper == COMPANY_ZETTRA.upper()
    
    # Check for customer-specific template files if customer info is provided
    # Use short customer name for lookup (from 'замовник')
    short_customer_name = None
    if source_workbook and customer_name:
        # Try to extract short name from full name if possible
        # If customer_name contains quotes, extract text inside quotes
        import re
        match = re.search(r'"([^"]+)"', customer_name)
        if match:
            short_customer_name = match.group(1)
        else:
            # If no quotes, use as is
            short_customer_name = customer_name
    
    if short_customer_name and source_workbook:
        customer_templates = get_customer_template_files(source_workbook, short_customer_name)
        if customer_templates:
            print(f"Customer '{short_customer_name}' has specific template files: {customer_templates}")
            
            # Determine which template file to use based on company
            template_base_dir = os.path.dirname(TEMPLATE_FILE)  # Same directory as default templates
            
            for template_file in customer_templates:
                template_path = os.path.join(template_base_dir, template_file)
                
                # Check if this template matches the company
                if is_avtotrans and ('zia' in template_file.lower() or 'зіа' in template_file.lower()):
                    if os.path.exists(template_path):
                        print(f"Using customer-specific {COMPANY_AVTOTRANS} template: {template_file}")
                        return template_path
                    else:
                        print(f"Warning: Customer template file not found: {template_path}")
                        
                elif is_zettra and ('zet' in template_file.lower() or 'зет' in template_file.lower() or 'template' in template_file.lower() or 'шаблон' in template_file.lower()) and 'zia' not in template_file.lower() and 'зіа' not in template_file.lower():
                    if os.path.exists(template_path):
                        print(f"Using customer-specific {COMPANY_ZETTRA} template: {template_file}")
                        return template_path
                    else:
                        print(f"Warning: Customer template file not found: {template_path}")
            
            print(f"No suitable customer-specific template found for company {company}, falling back to default")
    
    # Fall back to default templates
    if is_avtotrans:
        print(f"Using default {COMPANY_AVTOTRANS} template")
        return TEMPLATE_FILE
    elif is_zettra:
        print(f"Using default {COMPANY_ZETTRA} template")
        return TEMPLATE_FILE_ZETTRA
    else:
        print(f"WARNING: No template defined for company '{company}'")
        return None

def open_template(sort_by_excel_order=True, invoice_number=None):
    try:
        print("\n=== Starting template processing ===")
        result = get_visible_rows_data(invoice_number=invoice_number)
        if not result or len(result) != 7:
            print("\nERROR: Failed to get data from visible rows")
            return None
            
        companies_data, source_excel, source_wb, source_sheet, latest_date, source_columns, filtered_gid = result
        
        print(f"\nProcessing {len(companies_data)} rows:")
        for data in companies_data:
            print(f"Row {data['row']}: Company '{data['company']}', Processing row data...")
        
        # Sort rows by driver's last name alphabetically if not using Excel order
        if sort_by_excel_order:
            print("\n[REGISTRY] Using Excel table order for sorting")
        else:
            print("\n[REGISTRY] Using alphabetical order by driver's last name")
            # Sort alphabetically by driver's last name
            companies_data = sorted(companies_data, key=lambda row: extract_driver_last_name(row.get('водій', '')))
            print(f"Sorted driver order: {[extract_driver_last_name(row.get('водій', '')) for row in companies_data]}")
            
        # For now, let's work with the first visible row for template selection
        first_row_data = companies_data[0]
        print(f"\nUsing first visible row with company '{first_row_data['company']}'")
        
        is_kg_mode = False
        # Determine which template to use based on the company column
        customer_name = first_row_data.get('full_customer_name') or first_row_data.get('замовник', '')
        template_path = get_template_path(first_row_data['company'], customer_name, source_wb)
        if not template_path:
            print(f"ERROR: No template defined for company: {first_row_data['company']}")
            return None
            
        # Initialize Excel application for template
        print("\nOpening template...")
        template_excel = win32com.client.Dispatch("Excel.Application")
        template_excel.Visible = True  # Make Excel visible for debugging
        
        # Open template file
        if os.path.exists(template_path):
            template_wb = template_excel.Workbooks.Open(template_path)
            template_sheet = template_wb.Worksheets(1)  # Assume first sheet
            print(f"Successfully opened template: {template_path}")
            
            # Apply kg mode text replacements if needed
            if is_kg_mode:
                replace_template_text_kg_mode(template_wb)
                remove_vat_total_row(template_sheet)
            
            # Find and prepare table area
            table_info = find_table_range(template_sheet)
            if not table_info:
                return None
                
            # Adjust number of rows
            if not adjust_template_rows(template_sheet, table_info, len(companies_data), is_kg_mode):
                return None
            
            # Check if we need to hide loading weight column for specific customer
            customer_short_name = first_row_data.get('замовник', '')
            should_hide_loading_weight = customer_short_name.strip() == 'ТРІУМФ ФІД'
            
            if should_hide_loading_weight:
                print(f"\nCustomer is '{customer_short_name}' - hiding loading weight column")
                # Hide the loading weight column in the template
                hide_loading_weight_column(template_sheet, table_info, is_kg_mode)
                
            # Fill table data
            if not fill_table_data(template_sheet, table_info, companies_data, source_columns, is_kg_mode, should_hide_loading_weight):
                return None
            
            # Replace all template placeholders
            print("\nReplacing template placeholders:")
            
            # Handle customer name specially
            customer_name = first_row_data.get('full_customer_name') or first_row_data.get('замовник', '')
            if customer_name:
                replace_template_text(template_wb, 'customer', customer_name)
            
            # Handle company short name specially
            if customer_name:
                company_short_name = create_company_short_name(customer_name)
                replace_template_text(template_wb, 'companyshortname', company_short_name)
                print(f"Created company short name: '{company_short_name}' from '{customer_name}'")
            
            # Handle EDRPOU specially - lookup from Контрагенти sheet using short customer name
            customer_short_name = first_row_data.get('замовник', '')
            if customer_short_name:
                edrpou = get_customer_edrpou(source_wb, customer_short_name)
                if edrpou:
                    replace_template_text(template_wb, 'edrpou', edrpou)
                    print(f"Added EDRPOU: '{edrpou}' for customer '{customer_short_name}'")
                else:
                    print(f"No EDRPOU found for customer '{customer_short_name}'")
            
            # Handle lastdate / last date specially
            if latest_date:
                formatted_date = latest_date.strftime("%d.%m.%Y")
                replace_template_text(template_wb, 'lastdate', formatted_date)
                replace_template_text(template_wb, 'last date', formatted_date)
            
            # Handle other replacements
            for template_key, excel_col in COLUMN_MAPPINGS.items():
                if template_key not in ['customer', 'companyshortname', 'lastdate', 'last date', 'edrpou']:  # Skip already handled
                    value = first_row_data.get(excel_col, '')
                    # Always handle invoice field, even if empty
                    if template_key == 'invoice' or value:
                        replace_template_text(template_wb, template_key, value)
            
            # Determine template suffix for output fileformat
            template_suffix = Path(template_path).suffix.replace('.', '') if template_path else 'xlsx'
            
            # Generate save path using the latest date
            save_path = get_save_path(None, None, 
                                    first_row_data['замовник'], latest_date, filtered_gid=filtered_gid,
                                    invoice_number=invoice_number or first_row_data.get('рахунок'),
                                    company=first_row_data.get('company'),
                                    fileformat=template_suffix)
            
            return {
                'source': {'excel': source_excel, 'workbook': source_wb, 'sheet': source_sheet},
                'template': {'excel': template_excel, 'workbook': template_wb},
                'rows_data': companies_data,
                'save_path': save_path
            }
        else:
            print(f"ERROR: Template file not found at: {template_path}")
            return None
        
    except Exception as e:
        print(f"\nERROR in open_template: {str(e)}")
        return None

def validate_rows_consistency(rows_data):
    """Check if critical columns have the same values across all rows"""
    if not rows_data:
        return False, "No rows to validate"

    critical_columns = ['маршрут згідно товаро-транспортній накладній', 'вантаж', 'замовник']
    inconsistencies = []

    # Get values from first row as reference
    reference_values = {col: rows_data[0].get(col, '') for col in critical_columns}

    # Check each row against reference values
    for row_data in rows_data[1:]:
        for col in critical_columns:
            current_value = row_data.get(col, '')
            if current_value != reference_values[col]:
                inconsistencies.append(
                    f"Different values found in column '{col}':\n"
                    f"First row: '{reference_values[col]}'\n"
                    f"Row {row_data['row']}: '{current_value}'"
                )

    if inconsistencies:
        error_message = "\n".join([
            "WARNING: Inconsistencies found in selected rows:",
            *inconsistencies,
            "\nThis might indicate rows from different registries were selected."
        ])
        return False, error_message

    return True, "All rows are consistent"

def get_director_name(workbook, customer_name):
    return ""

def get_director_position(workbook, customer_name):
    return ""

def format_director_name(full_name):
    """Format director name as 'LastName F.M.'"""
    try:
        # Split the name into parts
        parts = full_name.strip().split()
        if len(parts) >= 3:
            last_name = parts[0]
            first_initial = parts[1][0]
            middle_initial = parts[2][0]
            return f"{last_name} {first_initial}.{middle_initial}."
        else:
            print(f"Warning: Could not format name '{full_name}' - insufficient parts")
            return full_name
    except Exception as e:
        print(f"Warning: Error formatting director name: {str(e)}")
        return full_name

def extract_driver_last_name(full_name):
    """Extract last name from a full driver name like 'Добролєжа Євген Ігорович'."""
    if not full_name or str(full_name).strip() == '':
        return ""
    
    # Split by spaces and take the first part (last name)
    parts = str(full_name).strip().split()
    if parts:
        return parts[0]
    return ""

def format_driver_name(full_name):
    """Format driver name from 'Добролєжа Євген Ігорович' to 'Добролєжа Є.І.'"""
    try:
        if not full_name or str(full_name).strip() == '':
            return ''
        
        # Split the name into parts
        parts = str(full_name).strip().split()
        if len(parts) >= 3:
            last_name = parts[0]
            first_initial = parts[1][0] + '.' if parts[1] else ''
            middle_initial = parts[2][0] + '.' if parts[2] else ''
            formatted_name = f"{last_name} {first_initial}{middle_initial}"
            print(f"Formatted driver name: '{full_name}' -> '{formatted_name}'")
            return formatted_name
        elif len(parts) == 2:
            # Handle case with only first and last name
            last_name = parts[0]
            first_initial = parts[1][0] + '.' if parts[1] else ''
            formatted_name = f"{last_name} {first_initial}"
            print(f"Formatted driver name (2 parts): '{full_name}' -> '{formatted_name}'")
            return formatted_name
        else:
            # If less than 2 parts, return as is
            print(f"Driver name too short to format: '{full_name}'")
            return full_name
    except Exception as e:
        print(f"Warning: Error formatting driver name '{full_name}': {str(e)}")
        return full_name

def validate_signature_data(customer_name, source_workbook):
    return []

def handle_customer_signature(sheet, customer_name=None, source_workbook=None):
    remove_signature_rows(sheet)

def remove_signature_rows(sheet):
    """Remove signature rows from the template"""
    try:
        # Find the cell containing 'Від перевізника'
        found = sheet.Cells.Find('Від перевізника')
        if found:
            # Delete 4 rows starting from the found cell's row
            start_row = found.Row
            sheet.Rows(f"{start_row}:{start_row + 3}").Delete()
            print("\nSignature rows removed successfully")
            return True
        else:
            print("\nWarning: Could not find 'Від перевізника' text, skipping signature removal")
            return False
    except Exception as e:
        print(f"\nWarning: Could not remove signature rows: {str(e)}")
        return False

def get_proposed_invoice_from_excel(last_proposed=None):
    return None

def get_invoice_input():
    """
    Reads invoice number from console.
    """
    print("")
    try:
        typed_input = input("Введіть номер рахунку (або введіть 'exit' для виходу): ").strip()
        return typed_input
    except (KeyboardInterrupt, EOFError):
        return 'exit'

def check_active_workbook_open():
    """Verify that the source file exists on disk."""
    global SOURCE_FILE
    if not SOURCE_FILE or not os.path.exists(SOURCE_FILE):
        # Try to find source file dynamically
        source_files = glob.glob(os.path.join(ROOT_DIR, "*.xls*"))
        source_files = [f for f in source_files if not os.path.basename(f).startswith('~$')]
        if len(source_files) == 1:
            SOURCE_FILE = source_files[0]
        else:
            print(f"\n[ERROR] Source file not found or multiple source files found in {ROOT_DIR}")
            print("Please make sure the source file (e.g. дебеторка.xlsx) is in the root directory.\n")
            sys.exit(1)
    return True

if __name__ == "__main__":
    import sys
    
    # Check if the workbook exists
    check_active_workbook_open()
    
    invoices = sys.argv[1:]
    
    if invoices:
        for inv in invoices:
            print(f"\n{'='*50}")
            print(f"PROCESSING INVOICE: {inv}")
            print(f"{'='*50}\n")
            
            result = open_template(sort_by_excel_order=True, invoice_number=inv)
            if result:
                # Save the template
                result['template']['workbook'].SaveAs(result['save_path']['gid_path'])
                print(f"\nFile saved successfully at: {result['save_path']['gid_path']}")
                print("Keeping the created workbook open in Excel.")
            else:
                print(f"Failed to process invoice {inv}")
    else:
        print("=== Registry Creation Helper ===")
        while True:
            try:
                inv = get_invoice_input()
                if not inv or inv.lower() == 'exit':
                    print("Exiting...")
                    break
                    
                print(f"\n{'='*50}")
                print(f"PROCESSING INVOICE: {inv}")
                print(f"{'='*50}\n")
                
                result = open_template(sort_by_excel_order=True, invoice_number=inv)
                if result:
                    # Save the template
                    result['template']['workbook'].SaveAs(result['save_path']['gid_path'])
                    print(f"\nFile saved successfully at: {result['save_path']['gid_path']}")
                    print("Keeping the created workbook open in Excel.")
                else:
                    print(f"Failed to process invoice {inv}")
            except (KeyboardInterrupt, EOFError):
                print("\nExiting...")
                break
